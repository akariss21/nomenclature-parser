import re
from flask import Flask, request, send_file, render_template
import os
import tempfile
import subprocess
from openpyxl import Workbook
from docx import Document

app = Flask(__name__)

@app.route('/upload')
def upload_page():
    return render_template('upload_page.html')

def process_column(column):
    results = []
    current_name = None
    rows_with_letters = []

    for index, cell in enumerate(column, start=2):
        cell = cell.strip()

        if not cell:
            continue

        if any(char.isalpha() for char in cell):
            name = re.sub(r'\d+', '', cell).strip()
            if name:
                current_name = name
                results.append([current_name])
                rows_with_letters.append(index)
    return results, rows_with_letters

def convert_index_to_number(index_str):
    match = re.match(r'\d+-(\d+)', index_str)
    if match:
        return f"{int(match.group(1)):02d}"
    return None

@app.route('/convert-to-excel', methods=['POST'])
def convert_to_excel():
    if 'document' not in request.files:
        return "Файл не найден в запросе", 400

    doc_file = request.files['document']
    if doc_file.filename == '':
        return "Файл не выбран", 400

    file_extension = os.path.splitext(doc_file.filename)[1].lower()
    if file_extension not in ['.docx']:
        return "Неподдерживаемый формат файла. Ожидаются или .docx", 400

    temp_doc = tempfile.NamedTemporaryFile(delete=False, suffix=file_extension)
    doc_file.save(temp_doc.name)

    content = []

    if file_extension == '.docx':
        try:
            doc = Document(temp_doc.name)
            for table in doc.tables:
                for i, row in enumerate(table.rows):
                    if i == 0:
                        continue

                    row_data = [cell.text.strip() for cell in row.cells]

                    if all(re.match(r'^\d+$', cell) for cell in row_data):
                        continue

                    content.append(row_data)
        except Exception as e:
            return f"Ошибка обработки файла .docx: {str(e)}", 500

    os.unlink(temp_doc.name)

    if not content:
        return "Таблица или текст не найдены в файле", 400

    column_1 = [row[0] for row in content if len(row) > 0]
    column_2 = [row[1] for row in content if len(row) > 1]

    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Extracted Data"

        headers = ["Вид документа","Индекс дела", "Заголовок дела", "Структурное подразделение/должность", "Литеры", "Мод регистрации", "Показать во всех СП", "Журнал регистрации",
                   "Тип хранения", "Срок хранения", "Номер пункта по типовому перечню", "Примечание"]
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        for i, row in enumerate(content, start=2):
            for j, value in enumerate(row, start=1):
                if j != 4 and j != 5:
                    sheet.cell(row=i, column=j, value=value)

            if len(row) > 5:
                column_4 = row[5]
                sheet.cell(row=i, column=11, value=column_4)
                if re.search(r'(постоянно|ДМН|ДЗН)', column_4, re.IGNORECASE):
                    sheet.cell(row=i, column=9, value="Постоянное")
                elif match := re.search(r'(\d+)\s*(год(?:а|ов)?|лет)', column_4, re.IGNORECASE):
                    years = match.group(1)
                    sheet.cell(row=i, column=10, value=years)
                    sheet.cell(row=i, column=9, value="Временное")

        processed_column_1, rows_with_letters = process_column(column_1)

        current_name = None
        current_index = 0
        for row_num in range(2, sheet.max_row + 1):
            if row_num in rows_with_letters:
                current_name = processed_column_1[current_index][0]
                current_index += 1

            if current_name:
                sheet.cell(row=row_num, column=4, value=current_name)

        for i, value in enumerate(column_1, start=2):
            sheet.cell(row=i, column=2, value=convert_index_to_number(value))

        for i, value in enumerate(column_2, start=2):
            sheet.cell(row=i, column=3, value=value)

        for i in range(2, sheet.max_row + 1):
            sheet.cell(row=i, column=1).value = None
        for row_num in sorted(rows_with_letters, reverse=True):
            for col_num in range(1, sheet.max_column + 1):
                sheet.cell(row=row_num, column=col_num).value = None
            sheet.delete_rows(row_num)
        workbook.save(temp_excel.name)
    except Exception as e:
        return f"Ошибка создания Excel: {str(e)}", 500

    response = send_file(temp_excel.name, as_attachment=True, download_name="converted.xlsx")
    os.unlink(temp_excel.name)
    return response

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=9999)
